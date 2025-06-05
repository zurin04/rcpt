import os
import logging
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file, session
from flask_sqlalchemy import SQLAlchemy
from flask_migrate import Migrate
from sqlalchemy.orm import DeclarativeBase
from werkzeug.middleware.proxy_fix import ProxyFix
from functools import wraps
from datetime import datetime, date, timedelta
import pytz
from sqlalchemy import func, and_
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from io import BytesIO

# Function removed as requested

# Set up logging
logging.basicConfig(level=logging.DEBUG)

class Base(DeclarativeBase):
    pass

db = SQLAlchemy(model_class=Base)

# Create the Flask app
app = Flask(__name__)
app.secret_key = os.environ.get("SESSION_SECRET", "default_secret_key_for_local_dev")
app.wsgi_app = ProxyFix(app.wsgi_app, x_proto=1, x_host=1)

# Configure the database
database_url = os.environ.get("DATABASE_URL")
if database_url and database_url.startswith("postgres://"):
    database_url = database_url.replace("postgres://", "postgresql://", 1)
app.config["SQLALCHEMY_DATABASE_URI"] = database_url or "sqlite:///receipts.db"
app.config["SQLALCHEMY_ENGINE_OPTIONS"] = {
    "pool_recycle": 300,
    "pool_pre_ping": True,
}

# Initialize the app with the extension
db.init_app(app)
migrate = Migrate(app, db)

# Define models here
class BusinessInfo(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    business_name = db.Column(db.String(200), nullable=False)
    business_email = db.Column(db.String(200))
    contact_number = db.Column(db.String(50), nullable=False)
    location = db.Column(db.String(200), nullable=False)
    attendant = db.Column(db.String(100), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    def __repr__(self):
        return f'<BusinessInfo {self.business_name}>'

class Receipt(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    receipt_number = db.Column(db.String(50), unique=True, nullable=False)
    date_created = db.Column(db.DateTime, default=datetime.now)
    
    # Business Information
    business_name = db.Column(db.String(200), nullable=False)
    business_email = db.Column(db.String(200))
    contact_number = db.Column(db.String(50), nullable=False)
    location = db.Column(db.String(200), nullable=False)
    attendant = db.Column(db.String(100), nullable=False)
    
    # Customer Information
    customer_name = db.Column(db.String(100))
    customer_address = db.Column(db.Text)
    
    # Receipt totals and payment
    total_amount = db.Column(db.Numeric(10, 2), default=0)
    money_received = db.Column(db.Numeric(10, 2), default=0)
    change_amount = db.Column(db.Numeric(10, 2), default=0)
    
    # Relationship to items
    items = db.relationship('ReceiptItem', backref='receipt', lazy=True, cascade='all, delete-orphan')

    def __init__(self, **kwargs):
        super(Receipt, self).__init__(**kwargs)
        if not self.receipt_number:
            self.receipt_number = datetime.now().strftime('%Y%m%d%H%M%S')
    
    def __repr__(self):
        return f'<Receipt {self.receipt_number}>'

class ReceiptItem(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    receipt_id = db.Column(db.Integer, db.ForeignKey('receipt.id'), nullable=False)
    
    description = db.Column(db.String(500), nullable=False)
    quantity = db.Column(db.Integer, nullable=False)
    price = db.Column(db.Numeric(10, 2), nullable=False)
    subtotal = db.Column(db.Numeric(10, 2), nullable=False)
    
    def __repr__(self):
        return f'<ReceiptItem {self.description}>'

# Admin authentication credentials
ADMIN_USERNAME = "admin"
ADMIN_PASSWORD = "dikosasabihin"

# Authentication decorator
def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('admin_logged_in'):
            return redirect(url_for('admin_login'))
        return f(*args, **kwargs)
    return decorated_function

@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    """Admin login page"""
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        if username == ADMIN_USERNAME and password == ADMIN_PASSWORD:
            session['admin_logged_in'] = True
            flash('Successfully logged in!', 'success')
            return redirect(url_for('admin_panel'))
        else:
            flash('Invalid username or password.', 'error')
    
    return render_template('admin_login.html')

@app.route('/admin/logout')
def admin_logout():
    """Admin logout"""
    session.pop('admin_logged_in', None)
    flash('You have been logged out.', 'info')
    return redirect(url_for('index'))

@app.route('/')
def index():
    """Main page with the receipt form"""
    # Get saved business information
    business_info = BusinessInfo.query.first()
    return render_template('form.html', business_info=business_info)

@app.route('/api/business_info', methods=['GET', 'POST'])
def business_info_api():
    """API for business information"""
    if request.method == 'POST':
        data = request.get_json()
        
        # Get or create business info (assuming single business)
        business_info = BusinessInfo.query.first()
        if not business_info:
            business_info = BusinessInfo()
            db.session.add(business_info)
        
        # Update fields
        business_info.business_name = data.get('business_name', '')
        business_info.contact_number = data.get('contact_number', '')
        business_info.location = data.get('location', '')
        business_info.business_email = data.get('business_email', '')
        business_info.attendant = data.get('attendant', '')
        
        try:
            db.session.commit()
            return jsonify({'success': True, 'message': 'Business information saved successfully'})
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'message': str(e)}), 500
    
    else:  # GET
        business_info = BusinessInfo.query.first()
        if business_info:
            return jsonify({
                'business_name': business_info.business_name,
                'contact_number': business_info.contact_number,
                'location': business_info.location,
                'business_email': business_info.business_email or '',
                'attendant': business_info.attendant
            })
        return jsonify({})

@app.route('/generate_receipt', methods=['POST'])
def generate_receipt():
    """Generate and display the receipt"""
    try:
        # Get form data
        business_name = request.form.get('business_name', '').strip()
        contact_number = request.form.get('contact_number', '').strip()
        location = request.form.get('location', '').strip()
        attendant = request.form.get('attendant', '').strip()
        
        # Optional business email
        business_email = request.form.get('business_email', '').strip()
        
        # Optional customer data
        customer_name = request.form.get('customer_name', '').strip()
        customer_address = request.form.get('customer_address', '').strip()
        
        # Payment data
        money_received = request.form.get('money_received', '0').strip()
        
        # Item data - handle multiple items
        item_descriptions = request.form.getlist('item_description[]')
        custom_descriptions = request.form.getlist('custom_description[]')
        quantities = request.form.getlist('quantity[]')
        prices = request.form.getlist('price[]')
        
        # Debug logging
        app.logger.debug(f"Item descriptions: {item_descriptions}")
        app.logger.debug(f"Quantities: {quantities}")
        app.logger.debug(f"Prices: {prices}")
        
        # Ensure we have lists
        if not isinstance(item_descriptions, list):
            item_descriptions = [item_descriptions] if item_descriptions else []
        if not isinstance(custom_descriptions, list):
            custom_descriptions = [custom_descriptions] if custom_descriptions else []
        if not isinstance(quantities, list):
            quantities = [quantities] if quantities else []
        if not isinstance(prices, list):
            prices = [prices] if prices else []
        
        # Validate required fields
        if not all([business_name, contact_number, location, attendant]):
            flash('Please fill in all required business information fields.', 'error')
            return redirect(url_for('index'))
        
        # Create receipt record
        receipt = Receipt(
            business_name=business_name,
            business_email=business_email,
            contact_number=contact_number,
            location=location,
            attendant=attendant,
            customer_name=customer_name,
            customer_address=customer_address
        )
        
        # Process items
        items = []
        total = 0
        
        for i in range(len(item_descriptions)):
            if i < len(quantities) and i < len(prices):
                desc = item_descriptions[i].strip()
                # Use custom description if main description is "custom"
                if desc == "custom" and i < len(custom_descriptions):
                    desc = custom_descriptions[i].strip()
                
                qty_str = quantities[i].strip()
                price_str = prices[i].strip()
                
                if desc and qty_str and price_str:
                    try:
                        qty = int(qty_str)
                        price = float(price_str)
                        subtotal = qty * price
                        total += subtotal
                        
                        # Create receipt item
                        receipt_item = ReceiptItem()
                        receipt_item.description = desc
                        receipt_item.quantity = qty
                        receipt_item.price = price
                        receipt_item.subtotal = subtotal
                        receipt.items.append(receipt_item)
                        
                        items.append({
                            'description': desc,
                            'quantity': qty,
                            'price': price,
                            'subtotal': subtotal
                        })
                    except (ValueError, TypeError):
                        flash(f'Invalid quantity or price for item: {desc}', 'error')
                        return redirect(url_for('index'))
        
        # Set total and payment information
        receipt.total_amount = total
        
        # Process payment information
        try:
            money_received_float = float(money_received) if money_received else 0
            receipt.money_received = money_received_float
            receipt.change_amount = money_received_float - float(total)
        except (ValueError, TypeError):
            receipt.money_received = 0
            receipt.change_amount = 0
        
        # Save to database
        try:
            db.session.add(receipt)
            db.session.commit()
            app.logger.info(f"Receipt {receipt.receipt_number} saved to database")
        except Exception as e:
            db.session.rollback()
            app.logger.error(f"Database error: {str(e)}")
            flash('Error saving receipt to database.', 'error')
            return redirect(url_for('index'))
        
        # Generate receipt data for template
        receipt_data = {
            'receipt_number': receipt.receipt_number,
            'date': receipt.date_created.strftime('%B %d, %Y'),
            'time': receipt.date_created.strftime('%I:%M %p'),
            'business_name': receipt.business_name,
            'business_email': receipt.business_email,
            'contact_number': receipt.contact_number,
            'location': receipt.location,
            'attendant': receipt.attendant,
            'customer_name': receipt.customer_name,
            'customer_address': receipt.customer_address,
            'receipt_items': items,
            'total': float(receipt.total_amount)
        }
        
        # Create a receipt display object similar to view_receipt for consistency
        class ReceiptDisplay:
            def __init__(self, receipt_obj):
                # Copy all attributes from the original receipt
                for attr in dir(receipt_obj):
                    if not attr.startswith('_') and not callable(getattr(receipt_obj, attr)):
                        setattr(self, attr, getattr(receipt_obj, attr))
                
                # Add formatted date and time using local timezone
                local_tz = pytz.timezone('Asia/Manila')  # Philippine timezone
                local_time = receipt_obj.date_created.replace(tzinfo=pytz.UTC).astimezone(local_tz)
                self.date = local_time.strftime('%B %d, %Y')
                self.time = local_time.strftime('%I:%M %p')
                
                # Ensure items are properly formatted for template
                self.items = [{'description': item.description, 'quantity': item.quantity, 
                              'price': float(item.price), 'subtotal': float(item.subtotal)} 
                             for item in receipt_obj.items]
                
                # Keep the original items relationship for template compatibility
                self.receipt_items = self.items
        
        receipt_display = ReceiptDisplay(receipt)
        return render_template('receipt.html', receipt=receipt_display)
        
    except Exception as e:
        app.logger.error(f"Error generating receipt: {str(e)}")
        import traceback
        app.logger.error(f"Full traceback: {traceback.format_exc()}")
        flash('An error occurred while generating the receipt. Please try again.', 'error')
        return redirect(url_for('index'))

@app.route('/receipts')
def receipts_list():
    """List all receipts"""
    receipts = Receipt.query.order_by(Receipt.date_created.desc()).all()
    return render_template('receipts_list.html', receipts=receipts)

@app.route('/receipt/<receipt_number>')
def view_receipt(receipt_number):
    """View a specific receipt"""
    receipt = Receipt.query.filter_by(receipt_number=receipt_number).first_or_404()
    
    # Create a receipt object that has both the original receipt data and properly formatted items
    class ReceiptDisplay:
        def __init__(self, receipt_obj):
            # Copy all attributes from the original receipt
            for attr in dir(receipt_obj):
                if not attr.startswith('_') and not callable(getattr(receipt_obj, attr)):
                    setattr(self, attr, getattr(receipt_obj, attr))
            
            # Add formatted date and time using local timezone
            local_tz = pytz.timezone('Asia/Manila')  # Philippine timezone
            local_time = receipt_obj.date_created.replace(tzinfo=pytz.UTC).astimezone(local_tz)
            self.date = local_time.strftime('%B %d, %Y')
            self.time = local_time.strftime('%I:%M %p')
            
            # Ensure items are properly formatted for template
            self.items = [{'description': item.description, 'quantity': item.quantity, 
                          'price': float(item.price), 'subtotal': float(item.subtotal)} 
                         for item in receipt_obj.items]
            
            # Keep the original items relationship for template compatibility
            self.receipt_items = self.items
    
    receipt_display = ReceiptDisplay(receipt)
    return render_template('receipt.html', receipt=receipt_display)

@app.route('/export_excel')
def export_excel():
    """Export all receipts to Excel"""
    receipts = Receipt.query.order_by(Receipt.date_created.desc()).all()
    
    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Receipts Report"
    
    # Define styles
    header_font = Font(bold=True, size=12)
    header_alignment = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Headers
    headers = ['Receipt #', 'Date', 'Business Name', 'Customer', 'Attendant', 
               'Items', 'Total Amount', 'Total in Words']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # Data rows
    for row, receipt in enumerate(receipts, 2):
        ws.cell(row=row, column=1, value=receipt.receipt_number).border = border
        ws.cell(row=row, column=2, value=receipt.date_created.strftime('%Y-%m-%d %H:%M')).border = border
        ws.cell(row=row, column=3, value=receipt.business_name).border = border
        ws.cell(row=row, column=4, value=receipt.customer_name or 'Walk-in').border = border
        ws.cell(row=row, column=5, value=receipt.attendant).border = border
        
        # Items summary
        items_summary = '; '.join([f"{item.description} ({item.quantity}x)" for item in receipt.items])
        ws.cell(row=row, column=6, value=items_summary).border = border
        
        ws.cell(row=row, column=7, value=float(receipt.total_amount)).border = border
    
    # Adjust column widths
    for col in range(1, 8):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20
    
    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    # Generate filename with current date
    filename = f"receipts_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return send_file(excel_file, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/print_all_receipts')
def print_all_receipts():
    """Print all receipts"""
    receipts = Receipt.query.order_by(Receipt.date_created.desc()).all()
    
    receipts_data = []
    for receipt in receipts:
        receipt_data = {
            'receipt_number': receipt.receipt_number,
            'date': receipt.date_created.strftime('%B %d, %Y'),
            'time': receipt.date_created.strftime('%I:%M %p'),
            'business_name': receipt.business_name,
            'business_email': receipt.business_email,
            'contact_number': receipt.contact_number,
            'location': receipt.location,
            'attendant': receipt.attendant,
            'customer_name': receipt.customer_name,
            'customer_address': receipt.customer_address,
            'receipt_items': [{'description': item.description, 'quantity': item.quantity, 
                              'price': float(item.price), 'subtotal': float(item.subtotal)} for item in receipt.items],
            'total': float(receipt.total_amount)
        }
        receipts_data.append(receipt_data)
    
    return render_template('print_all_receipts.html', receipts=receipts_data)

@app.route('/new_receipt')
def new_receipt():
    """Create a new receipt"""
    return render_template('form.html', clear_customer_data=True)

@app.route('/admin')
@admin_required
def admin_panel():
    """Business administration panel"""
    business_info = BusinessInfo.query.first()
    total_receipts = Receipt.query.count()
    recent_receipts = Receipt.query.order_by(Receipt.date_created.desc()).limit(5).all()
    
    # Calculate daily sales
    today = date.today()
    
    # Today's sales
    today_sales = db.session.query(func.sum(Receipt.total_amount)).filter(
        func.date(Receipt.date_created) == today
    ).scalar() or 0
    
    # Yesterday's sales
    yesterday = today - timedelta(days=1)
    yesterday_sales = db.session.query(func.sum(Receipt.total_amount)).filter(
        func.date(Receipt.date_created) == yesterday
    ).scalar() or 0
    
    # This week's sales (current week)
    week_start = today - timedelta(days=today.weekday())
    week_sales = db.session.query(func.sum(Receipt.total_amount)).filter(
        func.date(Receipt.date_created) >= week_start
    ).scalar() or 0
    
    # This month's sales
    month_start = today.replace(day=1)
    month_sales = db.session.query(func.sum(Receipt.total_amount)).filter(
        func.date(Receipt.date_created) >= month_start
    ).scalar() or 0
    
    # Daily sales for the last 7 days
    daily_sales = []
    for i in range(6, -1, -1):  # Last 7 days
        day = today - timedelta(days=i)
        day_total = db.session.query(func.sum(Receipt.total_amount)).filter(
            func.date(Receipt.date_created) == day
        ).scalar() or 0
        
        daily_sales.append({
            'date': day.strftime('%Y-%m-%d'),
            'day_name': day.strftime('%A'),
            'total': float(day_total),
            'receipts_count': Receipt.query.filter(func.date(Receipt.date_created) == day).count()
        })
    
    return render_template('admin.html', 
                         business_info=business_info, 
                         total_receipts=total_receipts, 
                         recent_receipts=recent_receipts,
                         today_sales=float(today_sales),
                         yesterday_sales=float(yesterday_sales),
                         week_sales=float(week_sales),
                         month_sales=float(month_sales),
                         daily_sales=daily_sales)

@app.route('/admin/daily-sales')
@admin_required
def daily_sales():
    """Daily sales tracking and analytics"""
    # Get date range from query parameters (default to last 30 days)
    end_date = date.today()
    start_date = end_date - timedelta(days=29)  # Last 30 days
    
    # Calculate daily sales for the date range
    daily_data = []
    current_date = start_date
    
    while current_date <= end_date:
        day_total = db.session.query(func.sum(Receipt.total_amount)).filter(
            func.date(Receipt.date_created) == current_date
        ).scalar() or 0
        
        receipts_count = Receipt.query.filter(
            func.date(Receipt.date_created) == current_date
        ).count()
        
        daily_data.append({
            'date': current_date.strftime('%Y-%m-%d'),
            'day_name': current_date.strftime('%A'),
            'day_short': current_date.strftime('%a'),
            'day_month': current_date.strftime('%d'),
            'total': float(day_total),
            'receipts_count': receipts_count
        })
        
        current_date += timedelta(days=1)
    
    # Calculate summary statistics
    total_sales = sum(day['total'] for day in daily_data)
    avg_daily_sales = total_sales / len(daily_data) if daily_data else 0
    best_day = max(daily_data, key=lambda x: x['total']) if daily_data else None
    total_receipts = sum(day['receipts_count'] for day in daily_data)
    
    return render_template('daily_sales.html',
                         daily_data=daily_data,
                         total_sales=total_sales,
                         avg_daily_sales=avg_daily_sales,
                         best_day=best_day,
                         total_receipts=total_receipts,
                         start_date=start_date.strftime('%Y-%m-%d'),
                         end_date=end_date.strftime('%Y-%m-%d'))

@app.route('/admin/print-daily-report')
@admin_required
def print_daily_report():
    """Print daily sales report"""
    today = date.today()
    
    # Get today's receipts with items
    today_receipts = Receipt.query.filter(
        func.date(Receipt.date_created) == today
    ).order_by(Receipt.date_created.desc()).all()
    
    # Calculate totals
    total_sales = sum(receipt.total_amount for receipt in today_receipts)
    total_receipts_count = len(today_receipts)
    
    # Get all items sold today with quantities
    items_summary = {}
    for receipt in today_receipts:
        for item in receipt.items:
            if item.description in items_summary:
                items_summary[item.description]['quantity'] += item.quantity
                items_summary[item.description]['total_amount'] += item.subtotal
            else:
                items_summary[item.description] = {
                    'quantity': item.quantity,
                    'total_amount': item.subtotal,
                    'price': item.price
                }
    
    # Convert to list and sort by quantity
    items_list = [
        {
            'description': desc,
            'quantity': data['quantity'],
            'price': data['price'],
            'total_amount': data['total_amount']
        }
        for desc, data in items_summary.items()
    ]
    items_list.sort(key=lambda x: x['quantity'], reverse=True)
    
    business_info = BusinessInfo.query.first()
    
    return render_template('print_daily_report.html',
                         today=today,
                         today_receipts=today_receipts,
                         total_sales=total_sales,
                         total_receipts_count=total_receipts_count,
                         items_summary=items_list,
                         business_info=business_info)

@app.route('/admin/business', methods=['GET', 'POST'])
@admin_required
def admin_business():
    """Manage business information"""
    if request.method == 'POST':
        # Get or create business info (assuming single business)
        business_info = BusinessInfo.query.first()
        if not business_info:
            business_info = BusinessInfo()
            db.session.add(business_info)
        
        # Update fields
        business_info.business_name = request.form.get('business_name', '')
        business_info.contact_number = request.form.get('contact_number', '')
        business_info.location = request.form.get('location', '')
        business_info.business_email = request.form.get('business_email', '')
        business_info.attendant = request.form.get('attendant', '')
        
        try:
            db.session.commit()
            flash('Business information updated successfully!', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating business information: {str(e)}', 'error')
        
        return redirect(url_for('admin_business'))
    
    business_info = BusinessInfo.query.first()
    return render_template('admin_business.html', business_info=business_info)

@app.route('/delete_receipt/<receipt_number>', methods=['POST'])
def delete_receipt(receipt_number):
    """Delete a specific receipt"""
    try:
        receipt = Receipt.query.filter_by(receipt_number=receipt_number).first()
        if not receipt:
            flash('Receipt not found.', 'error')
            return redirect(url_for('receipts_list'))
        
        db.session.delete(receipt)
        db.session.commit()
        flash(f'Receipt #{receipt_number} has been deleted successfully.', 'success')
        
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Error deleting receipt {receipt_number}: {str(e)}")
        flash('An error occurred while deleting the receipt.', 'error')
    
    return redirect(url_for('receipts_list'))

# Initialize database tables
with app.app_context():
    db.create_all()

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)